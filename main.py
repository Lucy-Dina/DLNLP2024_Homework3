import jieba
import gensim.models as w2v
from gensim.models import Word2Vec
from sklearn.cluster import KMeans
import random
import re
import numpy as np
import os
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# ====== initial ======
Paragraphs = [] # Sampled paragraphs
Seg_novel = []  # All preprocessed paragraphs (Model training object)
Novel_Num = 16

# ====== read novels & preprocess (filter & segment) ======
path = "D:\\Desktop\\zzzzzz\\2024_spring\\Deep_NLP\\Homework3\\jyxstxtqj_downcc.com"
novel_names = os.listdir(path)
print("novel_names:", novel_names)

# punctuation & stopwords
with open('cn_punctuation.txt', "r", encoding='gbk', errors='ignore') as fp:
    extra_words = list()
    for line in fp.readlines():
        line = line.strip()   # Remove the line break at the end of each line
        extra_words.append(line)
    fp.close()
with open('cn_stopwords.txt', "r", encoding='gbk', errors='ignore') as fs:
    for line in fs.readlines():
        line = line.strip()
        extra_words.append(line)
    fs.close()

# ad & Non_CN
ad = ['本书来自www.cr173.com免费txt小说下载站', '更多更新免费电子书请关注www.cr173.com', '----〖新语丝电子文库(www.xys.org)〗', '新语丝电子文库',
          '\u3000', '\n', '。', '？', '！', '，', '；', '：', '、', '《', '》', '“', '”', '‘', '’', '［', '］', '....', '......',
          '『', '』', '（', '）', '…', '「', '」', '\ue41b', '＜', '＞', '+', '\x1a', '\ue42b']
Non_CN = u'[^a-zA-Z0-9\u4e00-\u9fa5]'

# filter & segment
novel_num = 0
for novel_name in novel_names:
    novel_path = path + '\\' + novel_name
    with open(novel_path, "r", encoding='gbk', errors='ignore') as novel:
        paras = []  # lines of each novel, waiting for sampling
        print("Now is {}...".format(novel_name))
        line = novel.readline()
        while line:
            para = []  # each line of the novel
            output = '' # each line of the novel
            # filter & segment
            for a in ad:
                line = line.replace(a, '')
            line = re.sub(Non_CN, '', line)
            line_tmp = line.strip()
            line_seg = jieba.lcut(line_tmp.strip(), cut_all=False) # precision mode
            for word in line_seg:
                if (word not in extra_words) and (not word.isspace()):
                    para.append(word)
                    output += word
                    output += " "
            # save paragraphs
            if len(para) != 0:
                paras.append(para)
            if len(str(output.strip())) != 0:
                # Seg_novel.append(str(output.strip()))
                Seg_novel.append(str(output.strip()).split())   # str
            line = novel.readline()

        # randomly select two paragraphs from the novel text
        random_paragraphs = random.sample(paras, 2)
        Paragraphs.extend(random_paragraphs)

    print("...{} is over".format(novel_name))
    novel_num += 1

if novel_num == Novel_Num:
    print("=" * 40)
    print("All novels loaded completed！")

# ====== Word2Vec ======
# ---word training---
model1 = w2v.Word2Vec(sentences=Seg_novel, vector_size=200, window=5, min_count=5, sg=0)    # CBOW
model1.save('CBOW.model')
model2 = w2v.Word2Vec(sentences=Seg_novel, vector_size=200, window=5, min_count=5, sg=1)    # skip-gram
model2.save('skip-gram.model')
# Load the trained model
# model = Word2Vec.load('CBOW.model')
model = Word2Vec.load('skip-gram.model')

# ---Calculate the semantic distance between word vectors---
content = "The semantic distance between word vectors"
ws.cell(row= 1, column=1, value=content)
KeyPeople = [["范蠡", "西施", "王道", "阿青"],
             ["乾隆", "陈家洛", "霍青桐", "喀丝丽"],
             ["石破天", "石中玉", "丁当", "白自在"],
             ["张无忌", "赵敏", "殷素素", "周芷若"],
             ["段誉", "虚竹", "慕容复", "王语嫣"],
             ["郭靖", "黄蓉", "杨康", "穆念慈"],
             ["李文秀", "苏普", "马家骏", "阿曼"],
             ["袁承志", "安小慧", "温青青", "阿九"],
             ["杨过", "小龙女", "郭靖", "黄蓉"],
             ["令狐冲", "任我行", "岳灵珊", "东方不败"],
             ["阿青", "范蠡", "西施", "勾践"],
             ["狄云", "戚芳", "戚长发", "丁典"],
             ["胡一刀", "苗人凤", "苗若兰", "胡斐"],
             ["胡斐", "程灵素", "袁紫衣", "苗人凤"],
             ["林玉龙", "任飞燕", "常长风", "逍遥子"],
             ["康熙", "韦小宝", "陈近南", "方怡"]
             ]  # according to the rank of novel_names
for i in range(Novel_Num):
    content = "《{}》".format(novel_names[i])
    row_flag_distance = i + 2 + 3 * i   # i+1 -> row_flag_distance+4
    ws.cell(row=row_flag_distance, column=1, value=content)
    for j in range(4):
        if j < 3:
            word1 = KeyPeople[i][j]
            word2 = KeyPeople[i][j + 1]
            distance = model.wv.distance(word1, word2)
            print(f" In novel'{novel_names[i]}'，the semantic distance between'{word1}' and '{word2}' is: {distance}")
            content = "'{}' and '{}'".format(word1,word2)
            ws.cell(row=row_flag_distance + j, column=2, value=content)
            content = "{}".format(distance)
            ws.cell(row=row_flag_distance + j, column=3, value=content)
        if j == 3:
            word1 = KeyPeople[i][j]
            word2 = KeyPeople[i][0]
            distance = model.wv.distance(word1, word2)
            print(f" In novel'{novel_names[i]}'，the semantic distance between'{word1}' and '{word2}' is: {distance}")
            content = "'{}' and '{}'".format(word1, word2)
            ws.cell(row=row_flag_distance + j, column=2, value=content)
            content = "{}".format(distance)
            ws.cell(row=row_flag_distance + j, column=3, value=content)


# ---Clustering of a certain category of words---
content = "Clustering of a certain category of words"
row_flag_cluster = row_flag_distance + 4
ws.cell(row= row_flag_cluster, column=1, value=content)
# load word lists for different categories
with open('金庸小说全人物.txt', "r", encoding='gbk', errors='ignore') as fpn:
    people_names = list()
    for line in fpn.readlines():
        line = line.strip()
        jieba.add_word(line)
        people_names.append(line)
    fpn.close()
with open('金庸小说全武功.txt', "r", encoding='gbk', errors='ignore') as fkn:
    kungfu_names = list()
    for line in fkn.readlines():
        line = line.strip()
        jieba.add_word(line)
        kungfu_names.append(line)
    fkn.close()
with open('金庸小说全门派.txt', "r", encoding='gbk', errors='ignore') as fgn:
    gang_names = list()
    for line in fgn.readlines():
        line = line.strip()
        jieba.add_word(line)
        gang_names.append(line)
    fgn.close()
# verify whether a word exists in the trained model
people_list = []
gang_list = []
kungfu_list = []
for people in people_names:
    if people in model.wv.key_to_index: # Quickly search for the index of words or verify whether a word exists in the trained model
        people_list.append(people)
for gang in gang_names:
    if gang in model.wv.key_to_index:
        gang_list.append(gang)
for kungfu in kungfu_names:
    if kungfu in model.wv.key_to_index:
        kungfu_list.append(kungfu)
# print("people_list", people_list)
# print("gang_list", gang_list)
# print("kungfu_list", kungfu_list)
# Randomly select five to cluster
random_people = random.sample(people_list, 10)
random_gang = random.sample(gang_list, 10)
random_kungfu = random.sample(kungfu_list, 10)
print("random_people:", random_people)
print("random_gang:",random_gang)
print("random_kungfu:",random_kungfu)
word_list = random_people + random_gang + random_kungfu
vectors = [model.wv[word] for word in word_list]

# Using K-means clustering algorithm for clustering
k = 3  # Aggregated into three categories
kmeans = KMeans(n_clusters=k)
kmeans.fit(vectors)
j=1
for i, word in enumerate(word_list):    # Simultaneously obtaining element index and element value
    print(f"The {word} belongs to {kmeans.labels_[i]}")
    content = "'{}'".format(word)
    ws.cell(row=row_flag_cluster + j, column=1, value=content)
    content = "{}".format(kmeans.labels_[i])
    ws.cell(row=row_flag_cluster + j, column=2, value=content)
    j=j+1

# ---The semantic association of paragraphs---
content = "The semantic association of paragraphs"
row_flag_paragraphs = row_flag_cluster + j
ws.cell(row= row_flag_paragraphs, column=1, value=content)
# Calculate paragraph vectors (represented by average word vectors)
def get_paragraph_vector(paragraph, model):
    vectors = [model.wv[word] for word in paragraph if word in model.wv]
    if len(vectors) > 0:
        return np.mean(vectors, axis=0)
    else:
        return np.zeros(model.vector_size)

num = 0
flag = (2*Novel_Num)-1
h=1
while True:
    if num < flag:
        paragraph1 = Paragraphs[num]
        paragraph2 = Paragraphs[num + 1]
        vector1 = get_paragraph_vector(paragraph1, model)
        vector2 = get_paragraph_vector(paragraph2, model)
        # Calculate the similarity between paragraph vectors
        similarity = np.dot(vector1, vector2) / (np.linalg.norm(vector1) * np.linalg.norm(vector2)) # linear algebra norm
        print(f"The similarity between paragraph {num} and paragraph {num + 1} is: {similarity}")
        content = "Paragraph {} and Paragraph {}".format(num,num + 1)
        ws.cell(row=row_flag_paragraphs + h, column=1, value=content)
        content = "{}".format(similarity)
        ws.cell(row=row_flag_paragraphs + h, column=2, value=content)
        h = h + 1
        num = num + 1
    else:   # the last paragraph
        paragraph1 = Paragraphs[num]
        paragraph2 = Paragraphs[0]
        vector1 = get_paragraph_vector(paragraph1, model)
        vector2 = get_paragraph_vector(paragraph2, model)
        similarity = np.dot(vector1, vector2) / (np.linalg.norm(vector1) * np.linalg.norm(vector2))
        print(f"The similarity between paragraph {num} and paragraph 0 is: {similarity}")
        content = "Paragraph {} and Paragraph 0".format(num)
        ws.cell(row=row_flag_paragraphs + h, column=1, value=content)
        content = "{}".format(similarity)
        ws.cell(row=row_flag_cluster + h, column=2, value=content)
        h = h + 1
        break

# wb.save("outcome_CBOW.xlsx")
wb.save("outcome_skip-gram.xlsx")